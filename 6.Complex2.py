from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import time

path = "G:\\complexexcel.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active

driver = webdriver.Chrome()
driver.implicitly_wait(0.5)
driver.maximize_window()

#Copy 0 entries from the first column and google search

y = 2
while (y<12):
    driver.get('https://forms.gle/VQ8jgu78v7fLU5eA8')
    z = 1
    while (z<3):
        
        #NAME
        names = driver.find_element(By.XPATH, '//*[@id="mG61Hd"]/div[2]/div/div[2]/div[1]/div/div/div[2]/div/div[1]/div/div[1]/input')
        cellvalue = sheet.cell(y,z).value
        names.send_keys(cellvalue)
        #EMAIL
        z += 1
        emails = driver.find_element(By.XPATH, '//*[@id="mG61Hd"]/div[2]/div/div[2]/div[2]/div/div/div[2]/div/div[1]/div/div[1]/input')
        cellvalue2 = sheet.cell(y,z).value
        emails.send_keys(cellvalue2)
        #PHONENUMBER
        z += 1
        phnumber = driver.find_element(By.XPATH, '//*[@id="mG61Hd"]/div[2]/div/div[2]/div[3]/div/div/div[2]/div/div[1]/div/div[1]/input')
        cellvalue3 = sheet.cell(y,z).value
        phnumber.send_keys(cellvalue3)
        #SUBMIT
        driver.implicitly_wait(5000)
        time.sleep(3)
        button1 = driver.find_element(By.XPATH,'//*[@id="mG61Hd"]/div[2]/div/div[3]/div[1]/div[1]/div')
        button1.click()
    y+=1
    





